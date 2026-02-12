import sys # Easy way for Python to restart
import pulp  # Bring in the PuLP toolbox.
import time # Used to show how much time it takes to solve the rota
import calendar  # Bring in the calendar toolbox for month days.
import os # Bring in interaction with Windows/Apple/Linux
import json # Bring in functionality of saving/loading javascript object notation - data-interchange format
from tkinter import *  # Bring in the Tkinter toolbox for the window (GUI).
from tkinter import filedialog
import openpyxl # Allows to use .xlsx files
from selection_popups import prefer_count, cannot_count, manual_count # bring popup_select_shifts functions from a another file
from solver import solve_rota # bring PuLP solver from another file

# --------------------------------------------------------------------
# App plan:
# Part I: Tkinter GUI
# Part II: PuLP Solve
# --------------------------------------------------------------------

# --------------------------------------------------------------------
# Tkinter GUI
# --------------------------------------------------------------------

# Global variables: IMPORTANT!!!
year = None
month = None
holiday_days = [] # Empty variable, so "Make shifts" works even if Holidays saved nothing
shifts_list = []
workers_list = []
units_list = [] # Empty list to store units: "Cardiology", "Internal Medicine - Endocrinology" etc.
selected_cannot_days = {}  # Changed to dict to store per row_num
selected_prefer_days = {}  # Dict to store preferred days per row_num
selected_manual_days = {}  # Dict to store manual days per row_num

# Global variables: constant
day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]  # Group of day names.

# Settings for PuLP: points, hard rules; Other settings: shift making
points_filled = 100
points_preferred = 5
points_spacing = -1
spacing_days_threshold = 5  # How many days apart triggers the penalty
points_24hr = -10
enforce_no_adj_nights = True   # Night → Night next day
enforce_no_adj_days = True   # Day → Day next day hard rule
include_weekday_days = False   # False = default behaviour (skip Mon-Fri day shifts when making shifts)

# -------------------------------------------------------
# The main window (like the car's dashboard).
# -------------------------------------------------------

root = Tk()  # Make the window box.
#root.geometry("975x450")  # Set predetermined window size (width x height)
root.title("Hospital Shift Manager 'Riaukapp'")  # Name on top.

# Label inside frame.
Label(root, text="Enter the year (e.g., 2026):").pack()  # Text, side=LEFT for horizontal.

# Frame for year row – like a shelf for horizontal.
year_frame = Frame(root)  # Make small box (Frame) inside root.
year_frame.pack()  # Put the shelf on the window.

Label(root, text="Enter the month (1-12, e.g., 1 for January):").pack()  # Text label.

# Type box inside frame.
year_entry = Entry(year_frame)  # Type box in frame.
year_entry.pack(side=LEFT)  # Next to label.

# Function for save year (like button press).
def save_year():
    global year  # Use the year box outside.
    year_input = year_entry.get().strip()  # Get from type box and strip whitespace.
    try:
        year = int(year_input)
        if year < 1900 or year > 2100:
            error_label.config(text="Bad year – 1900-2100.")  # Show error.
        else:
            current_year_label.config(text="Current Year: " + str(year))  # Update show.
            error_label.config(text="")  # Clear error.
    except ValueError:
        error_label.config(text="Not a number!")

# Button inside frame.
Button(year_frame, text="Save", command=save_year).pack(side=LEFT)  # Button next.

# Show current year.
current_year_label = Label(year_frame, text="Current Year: None")  # Show label.
current_year_label.pack(side=LEFT)

month_frame = Frame(root)
month_frame.pack()

# Label and type box for month.
month_entry = Entry(month_frame)  # Type box.
month_entry.pack(side=LEFT)

def save_month():  # Function for the button.
    global month, num_days, starting_weekday, days_list, shifts_list  # Use these boxes outside – add shifts_list.
    month_input = month_entry.get()  # Get from type box.
    try:
        month = int(month_input)
        if month < 1 or month > 12:
            error_label.config(text="Please enter a month between 1 and 12.")  # Error.
            return  # Stop early if bad.
        # Calculate details.
        month_details = calendar.monthrange(year, month)
        starting_weekday = month_details[0]  # Weekday for Day 1.
        num_days = month_details[1]  # Days in month.
        # Make the days list
        days_list = []  # Empty list to hold days.
        for day in range(1, num_days + 1):  # Loop from 1 to num_days +1 (to include last).
            days_list.append(day)  # Add day to list.
        # Show the details.
        day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]  # Group of day names.
        # Show the days list.
        month_name_list = ["January", "February", "March", "April", "May", "June", "July", "September", "October", "November", "December"]
        month_name = month_name_list[month-1]
        current_month_label.config(text="Current Month: " + str(month_name)) # Update month
        error_label.config(text=f"{month_name} has {len(days_list)} days, start of the weekday: {day_names[starting_weekday]}")  # Update result.
    except ValueError:
        error_label.config(text="That's not a number! Please try again.")  # Error.

Button(month_frame, text="Save", command=save_month).pack(side=LEFT)  # Button, click runs save_month.

current_month_label = Label(month_frame, text="Current Month: None")
current_month_label.pack(side=LEFT)

Label(root, text="Units (comma-separated, e.g. Internal Medicine,Cardiology):").pack()

units_frame = Frame(root, borderwidth=2, relief="groove")
units_frame.pack()

units_entry = Entry(units_frame, width=40)
units_entry.pack(side=LEFT)

def save_units():
    global units_list
    units_input = units_entry.get().strip()
    if units_input == "":
        units_list = []
        current_units_label.config(text="Current Units: None")
        error_label.config(text="")
        return
    
    # Split by comma and clean up whitespace
    units_list = [u.strip() for u in units_input.split(",") if u.strip()]
    
    if not units_list:
        current_units_label.config(text="Current Units: None")
        error_label.config(text="Please enter at least one unit.")
        return
    
    current_units_label.config(text=f"Current Units: {', '.join(units_list)}")
    error_label.config(text=f"{len(units_list)} unit(s) saved.")

Button(units_frame, text="Save", command=save_units).pack(side=LEFT)

current_units_label = Label(units_frame, text="Current Units: None")
current_units_label.pack(side=LEFT)

# =======================================
# Popup functions, found in separate file
# =======================================

def show_prefer_popup(row_num):
    """
    Opens a popup window where the user can choose which shifts this worker
    would LIKE to work (preferred days/shifts).

    What it does, step by step:
    1. Creates a new small window with title "Select Days Prefer Work"
    2. Lists every day of the month + checkboxes for Day and Night shifts
    3. Pre-ticks boxes that were already selected before (if the user opened it again)
    4. Includes "Check All Day" and "Check All Night" boxes for fast selection
    5. When user clicks "Save Selection":
       - Saves chosen preferred shifts into selected_prefer_days[row_num]
       - Changes button text to show how many were selected (e.g. "Select (4)")
       - Shows message like "Worker prefers to work on these shifts: ..."
       - Closes the popup

    - Similar function for the other popups
    """

    prefer_popup_inputs = {
        "give_root": root,
        "give_days_list": days_list,
        "give_starting_weekday": starting_weekday,
        "give_worker_rows": worker_rows,
        "give_row_num": row_num,
        "give_selected_prefer_days": selected_prefer_days,   
        "give_error_label": error_label,                     
    }

    prefer_count(prefer_popup_inputs)


def show_cannot_popup(row_num):

    cannot_popup_inputs = {
        "give_root": root,
        "give_days_list": days_list,
        "give_starting_weekday": starting_weekday,
        "give_worker_rows": worker_rows,
        "give_row_num": row_num,
        "give_selected_cannot_days": selected_cannot_days,   
        "give_error_label": error_label,                
    }

    cannot_count(cannot_popup_inputs)

def show_manual_popup(row_num):

    manual_popup_inputs = {
        "give_root": root,
        "give_days_list": days_list,
        "give_starting_weekday": starting_weekday,
        "give_worker_rows": worker_rows,
        "give_row_num": row_num,
        "give_selected_manual_days": selected_manual_days,   
        "give_error_label": error_label,
        "give_units_list": units_list,    
    }
    
    manual_count(manual_popup_inputs)


Label(root, text="Enter public holiday days, separated by comma (e.g., 24,25,26) or leave blank:").pack()  # Text label.

holiday_frame = Frame(root)
holiday_frame.pack()

# Label and type box for holidays.
holiday_entry = Entry(holiday_frame)  # Type box.
holiday_entry.pack(side=LEFT)

def save_holidays():  # Function for the button.
    global holiday_days  # Use the holiday_days box outside.
    holiday_input = holiday_entry.get()  # Get from type box.
    holiday_days = []  # Start empty.
    if holiday_input.strip() == "":  # If blank.
        holiday_days = []
        holidays_label.config(text="Holiday List: None")
        error_label.config(text="")  # Clear error.
        return  # Done.
    parts = holiday_input.split(",")  # Cut at commas.
    try:
        holiday_days = [int(part.strip()) for part in parts]  # Turn to numbers.
        invalid_days = []
        for h_day in holiday_days[:]:  # Copy to avoid remove issues.
            if h_day < 1 or h_day > num_days:  # Out of range.
                invalid_days.append(h_day)
                holiday_days.remove(h_day)
        if invalid_days:
            error_label.config(text=f"Error: These days are invalid (must be 1-{num_days}): {invalid_days}")  # Show error.
            return  # Stop early.
        holidays_label.config(text="Holiday List: " + str(holiday_days))
        error_label.config(text="")  # Clear if good.
    except ValueError:
        error_label.config(text="Error: Some entries weren't numbers. Please try again.")  # Error.

Button(holiday_frame, text="Save", command=save_holidays).pack(side=LEFT)  # Button, click runs save_holidays.

holidays_label = Label(holiday_frame, text="Holiday List: None")
holidays_label.pack(side=RIGHT)

make_shifts_button_frame = Frame(root)
make_shifts_button_frame.pack()

add_worker_button_frame = Frame(root)
add_worker_button_frame.pack()

# Button to add row.

# Frame to hold the scrollable worker area (like a shelf for the canvas)
worker_container = Frame(root)
#worker_container.pack(fill="both", expand=True)  # Fill the space, can grow
worker_container.pack()

# Create the canvas (drawing board)
global worker_canvas, worker_scrollbar, worker_inner_frame
worker_canvas = Canvas(worker_container, width=830)  # Height=200 pixels – change if you want taller/shorter
worker_canvas.pack(side=LEFT, fill="both")  # Put on left, fill space

# Create the scrollbar and link it to the canvas
worker_scrollbar = Scrollbar(worker_container, orient="vertical", command=worker_canvas.yview)
worker_scrollbar.pack(side=RIGHT, fill="y")  # Put on right, vertical

# Link canvas back to scrollbar (so it knows when to show the slider)
worker_canvas.configure(yscrollcommand=worker_scrollbar.set)

# Create the inner frame (where worker rows will go)
worker_inner_frame = Frame(worker_canvas)
worker_canvas.create_window((0, 0), window=worker_inner_frame, anchor="nw")  # Put the frame at top-left of canvas

# Mouse wheel
def on_mouse_wheel(event):
    worker_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")  # Scroll up/down with wheel

root.bind("<MouseWheel>", on_mouse_wheel)  # Link wheel to function

# Make the canvas scroll when the inner frame grows
def update_scroll_region(event):
    worker_canvas.configure(scrollregion=worker_canvas.bbox("all"))  # Keep this – updates scroll area

    # Ask inner_frame how tall it needs to be (reqheight = "required height")
    required_height = worker_inner_frame.winfo_reqheight()  # This is like measuring the stack of rows

    # Decide the height: Min of required or 200 (the max)
    max_height = 300  # Your max – change if you want
    new_height = min(required_height, max_height)  # min() picks the smaller one

    # Set the canvas height to that
    worker_canvas.config(height=new_height)  # Update it!

    # If required > max, scrollbar will show automatically – no extra code needed

def update_inner_width(event):
    worker_inner_frame.config(width=event.width)  # Set width to match canvas

worker_canvas.bind("<Configure>", update_inner_width)  # Run this when canvas size changes

worker_inner_frame.bind("<Configure>", update_scroll_region)  # Run this function when inner frame changes size

def make_shifts():  # New function for making shifts list.
    global shifts_list, holiday_days, include_weekday_days, units_list  # Use the shifts_list box outside.

    if year == None or month == None:
        error_label.config(text="Erorr: Please select year and month first!")
        return

    # Check if units are defined
    if not units_list:
        error_label.config(text="Error: Please define units first!")
        return

    shifts_list = []  # Empty list for shifts.
    shift_types = ["Day", "Night"]  # Group of types.
    day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]  # Group of day names.
    
    # Loop through each unit
    for unit in units_list:
        for day in days_list:  # Loop each day.
            weekday = (starting_weekday + (day - 1)) % 7  # Calculate day name ( %7 like clock wrap).
            tags = [day_names[weekday]]  # Add day name tag.
            if weekday in [5, 6]:  # Sat/Sun – weekend.
                tags.append("Weekend")  # Add tag.
            if day in holiday_days:  # If holiday.
                tags.append("Public holiday")  # Add tag.
            for shift_type in shift_types:  # Loop Day then Night.
                # Check if this is a Monday-Friday Day shift (and not a holiday) – if both true, exclude (skip)
                if (not include_weekday_days) and weekday in [0, 1, 2, 3, 4] and shift_type == "Day" and day not in holiday_days:
                    continue  # Skip Mon-Fri day shifts only when the checkbox is OFF
                shift_name = f"{shift_type} {day} {unit}"  # Make name: ex. Cardiology Day 1, Internal Medicine Night 2...
                shift_dict = {  # Make the shift box.
                    "name": shift_name,
                    "type": shift_type,
                    "tags": tags,
                    "unit": unit,
                    "assigned_worker": None
                }
                shifts_list.append(shift_dict)  # Add to list.
        
        error_label.config(text=f"Shifts made: {len(shifts_list)} across {len(units_list)} unit(s)") # Update the label with count.
        
        # Show the shifts list nicely in terminal, for debugging
        print("Your shifts list (with tags, types):")
        for shift in shifts_list:  # Loop to print each one.
            print(f"Shift: {shift['name']}, Tags: {shift['tags']}, Type: {shift['type']}, Assigned: {shift['assigned_worker']}")

Button(make_shifts_button_frame, text="Make Shifts", command=make_shifts, width=10).pack()  # Button, click runs make_shifts.

# Box to show the shifts list.
#shifts_label = Label(root, text="Shifts made: None")  # Start None.
#shifts_label.pack()  # Put on window.

# Frame for workers – like a big shelf for the table.
workers_frame = Frame(worker_inner_frame)  # NEW: Pack to inner_frame
workers_frame.pack(fill="x")  # NEW: Fill horizontal, stack vertical

# Configure columns to have fixed width of 50
for col in range(10):
    workers_frame.columnconfigure(col, minsize=10, weight=0)

# Global headers for columns.
Label(workers_frame, text="Name", width=10, padx=2, pady=2).grid(row=0, column=0, sticky="ew", padx=2, pady=2)
Label(workers_frame, text="Shift Range", width=10, padx=2, pady=2).grid(row=0, column=1, sticky="ew", padx=2, pady=2)
Label(workers_frame, text="Cannot Days", width=10, padx=2, pady=2).grid(row=0, column=2, sticky="ew", padx=2, pady=2)
Label(workers_frame, text="Prefer Days", width=10, padx=2, pady=2).grid(row=0, column=3, sticky="ew", padx=2, pady=2)
Label(workers_frame, text="Max Wknds", width=10, padx=2, pady=2).grid(row=0, column=4, sticky="ew", padx=2, pady=2)
Label(workers_frame, text="Max 24hr", width=10, padx=2, pady=2).grid(row=0, column=5, sticky="ew", padx=2, pady=2)
Label(workers_frame, text="Manual Shifts", width=10, padx=2, pady=2).grid(row=0, column=6, sticky="ew", padx=2, pady=2)
Label(workers_frame, text="Save", width=10, padx=2, pady=2).grid(row=0, column=7, sticky="ew", padx=2, pady=2)
Label(workers_frame, text="Assign Manual", width=10, padx=2, pady=2).grid(row=0, column=8, sticky="ew", padx=2, pady=2)
Label(workers_frame, text="Delete", width=10, padx=2, pady=2).grid(row=0, column=9, sticky="ew", padx=2, pady=2)

# List to hold worker rows (for delete).
worker_rows = []  # Empty list to store widget references for each row.
worker_row_number = 1  # Start row 1 (after 0 headers)

def add_worker_row():  # Function for "Add Worker" button.
    global worker_row_number, workers_list
    row_num = worker_row_number  # Capture the current row number for this row

    # Column 0: Name box - placed directly in workers_frame.
    name_entry = Entry(workers_frame, width=10)  # Type box.
    name_entry.grid(row=row_num, column=0, sticky="ew", padx=2, pady=2)

    # Column 1: Shift range box.
    range_entry = Entry(workers_frame, width=10)
    range_entry.grid(row=row_num, column=1, sticky="ew", padx=2, pady=2)

    # Column 2: Cannot work button.
    cannot_button = Button(workers_frame, text="Select", width=10, command=lambda: show_cannot_popup(row_num))
    cannot_button.grid(row=row_num, column=2, sticky="ew", padx=2, pady=2)

    # Column 3: Prefer button.
    prefer_button = Button(workers_frame, text="Select", width=10, command=lambda: show_prefer_popup(row_num))
    prefer_button.grid(row=row_num, column=3, sticky="ew", padx=2, pady=2)

    # Column 4: Max weekends box.
    max_weekends_entry = Entry(workers_frame, width=10)
    max_weekends_entry.grid(row=row_num, column=4, sticky="ew", padx=2, pady=2)

    # Column 5: Max 24-hour box.
    max_24hr_entry = Entry(workers_frame, width=10)
    max_24hr_entry.grid(row=row_num, column=5, sticky="ew", padx=2, pady=2)

    # Column 6: Manual shifts button.
    manual_button = Button(workers_frame, text="Select", width=10, command=lambda: show_manual_popup(row_num))
    manual_button.grid(row=row_num, column=6, sticky="ew", padx=2, pady=2)

    # Column 7: Save Worker button.
    save_button = Button(workers_frame, text="Save", width=10, command=lambda: save_worker(row_num))
    save_button.grid(row=row_num, column=7, sticky="ew", padx=2, pady=2)

    # Column 8: Manual save button.
    manual_save_button = Button(workers_frame, text="Assign", width=10, command=lambda: save_manual(row_num))
    manual_save_button.grid(row=row_num, column=8, sticky="ew", padx=2, pady=2)

    # Column 9: Delete button.
    delete_button = Button(workers_frame, text="Delete", width=10, command=lambda: delete_row(row_num))
    delete_button.grid(row=row_num, column=9, sticky="ew", padx=2, pady=2)

    # Store all widgets for this row for deletion purposes
    row_widgets = {
        'name_entry': name_entry,
        'range_entry': range_entry,
        'cannot_button': cannot_button,
        'prefer_button': prefer_button,
        'max_weekends_entry': max_weekends_entry,
        'max_24hr_entry': max_24hr_entry,
        'manual_button': manual_button,
        'save_button': save_button,
        'manual_save_button': manual_save_button,
        'delete_button': delete_button,
        'row_num': row_num
    }
    worker_rows.append(row_widgets)
    print(f"Current row: {worker_row_number}") #Debugging
    worker_row_number += 1  # Add 1.

    def save_worker(row_num):  # The function – the "recipe" for saving a worker.
        # This runs when you click "Save Worker".
        # Get what you typed from the boxes.
        name = name_entry.get().strip() # Get the name
        range_input = range_entry.get().strip()  # Get shift range.
        max_weekends_input = max_weekends_entry.get().strip()  # Get max weekends.
        max_24hr_input = max_24hr_entry.get().strip()  # Get max 24-hour.

        # Check if shifts_list is created
        global shifts_list
        try:
            if not shifts_list:
                error_label.config(text="Shifts list is not created. Cannot save worker.")
                return
        except NameError:
            error_label.config(text="Shifts list is not created. Cannot save worker.")
            return

        # Unassign all shifts assigned to this worker's name before saving
        for shift in shifts_list:
            if shift["assigned_worker"] == name:
                shift["assigned_worker"] = None

        # Check if name is not blank.
        if name == "":  # If empty.
            error_label.config(text="Error: Name can't be blank.")  # Show error.
            return  # Stop early.

        # Check shift range (like "1-4").
        if range_input == "":  # If blank.
            error_label.config(text="Error: Shift range can't be blank.")  # Error.
            return
        range_parts = range_input.split("-")  # Cut at "-".
        if len(range_parts) != 2:  # Must be 2.
            error_label.config(text="Error: Range like 1-4.")  # Error.
            return
        try:
            min_shifts = int(range_parts[0])
            max_shifts = int(range_parts[1])
            if min_shifts > max_shifts or min_shifts < 0:
                error_label.config(text="Error: Min <= Max, positive.")  # Error.
                return
        except ValueError:
            error_label.config(text="Error: Range not numbers.")  # Error.
            return

        # Cannot work from selected days.
        global selected_cannot_days
        cannot_work = selected_cannot_days.get(row_num, [])  # Get the list for this row_num.

        prefer = selected_prefer_days.get(row_num, [])  # Get the list for this row_num.

        # Max weekends and 24hr – numbers.
        max_weekends = 100  # Default 100.
        if max_weekends_input != "":
            try:
                max_weekends = int(max_weekends_input)
                if max_weekends < 0:
                    error_label.config(text="Error: Max weekends positive or 0.")
                    return
            except ValueError:
                error_label.config(text="Error: Max weekends not number.")
                return

        max_24hr = 100  # Default 100.
        if max_24hr_input != "":
            try:
                max_24hr = int(max_24hr_input)
                if max_24hr < 0:
                    error_label.config(text="Error: Max 24hr positive or 0.")
                    return
            except ValueError:
                error_label.config(text="Error: Max 24hr not number.")
                return

        # ===== Check if worker already exists =====
        # Look through workers_list to find this worker by row number
        existing_worker = None  # Start with None (not found).
        for worker in workers_list:
            if worker["worker_row_number"] == row_num:
                existing_worker = worker  # Found them!
                break  # Stop looking.

        if existing_worker:  # If we found them (updating)
            # Update the existing worker's data
            existing_worker["name"] = name
            existing_worker["shifts_to_fill"] = [min_shifts, max_shifts]
            existing_worker["cannot_work"] = cannot_work
            existing_worker["prefers"] = prefer
            existing_worker["max_weekends"] = max_weekends
            existing_worker["max_24hr"] = max_24hr
            # Check if there were manually selected shifts
            had_manual_shifts = bool(selected_manual_days.get(row_num, []))
            message = f"Worker '{name}' updated!"
            if had_manual_shifts:
                message += " Manually assigned shifts cleared."
            error_label.config(text=message)  # Show success.
        else:  # If not found (new worker)
            # Create new worker dictionary
            worker_dict = {
                "name": name,
                "shifts_to_fill": [min_shifts, max_shifts],
                "cannot_work": cannot_work,
                "prefers": prefer,
                "max_weekends": max_weekends,
                "max_24hr": max_24hr,
                "worker_row_number": row_num
            }
            workers_list.append(worker_dict)  # Add new worker.
            error_label.config(text=f"Worker '{name}' saved!")  # Show success.

        # Print for debugging
        print("Your full worker list:")
        for worker in workers_list:
            print(f"Name: {worker['name']}, shifts: {worker['shifts_to_fill']}, Cannot: {worker['cannot_work']}, Prefers: {worker['prefers']}, Max weekends: {worker['max_weekends']}, Max 24hr: {worker['max_24hr']}, Row number: {worker['worker_row_number']}")

    def save_manual(row_num):
        # Get worker name
        name = name_entry.get().strip()
        if not name:
            error_label.config(text="Error: Worker name required for manual save.")
            return

        # Find worker
        worker = None
        for w in workers_list:
            if w.get("worker_row_number") == row_num:
                worker = w
                break
        if not worker:
            error_label.config(text="Error: Save worker first before manual assignment.")
            return

        # Get selected shifts
        manual_shifts = selected_manual_days.get(row_num, [])
        if not manual_shifts:
            error_label.config(text="No manual shifts selected.")
            return

        successful_assigns = 0
        for shift_name in manual_shifts:
            # Find shift in shifts_list
            shift = None
            for s in shifts_list:
                if s["name"] == shift_name:
                    shift = s
                    break
            if not shift:
                error_label.config(text=f"Error: Shift {shift_name} not found.")
                print("Error: Shift {shift_name} not found.")
                continue
            if shift["assigned_worker"] is not None:
                error_label.config(text=f"Error: Shift {shift_name} already assigned.")
                print("Error: Shift {shift_name} already assigned.")
                continue
            # Assign
            shift["assigned_worker"] = name
            successful_assigns += 1

        if successful_assigns > 0:
            # Update worker's range
            original_min = worker["shifts_to_fill"][0]
            original_max = worker["shifts_to_fill"][1]
            new_min = max(0, original_min - successful_assigns)
            new_max = max(0, original_max - successful_assigns)
            worker["shifts_to_fill"] = [new_min, new_max]
            range_entry.delete(0, END)
            range_entry.insert(0, f"{new_min}-{new_max}")
            error_label.config(text=f"Assigned {successful_assigns} manual shifts to {name}. Updated range to {new_min}-{new_max}.")
            for shift in shifts_list:
                print(shift)
            for worker in workers_list:
                print(worker)
        else:
            error_label.config(text="No shifts assigned.")


def delete_row(row_num):
    # Find the worker's name before deleting
    worker_name = None
    global workers_list, shifts_list, worker_rows
    for worker in workers_list:
        if worker.get("worker_row_number") == row_num:
            worker_name = worker["name"]
            break

    # Unassign all shifts assigned to this worker
    if worker_name:
        for shift in shifts_list:
            if shift["assigned_worker"] == worker_name:
                shift["assigned_worker"] = None

    # Clean up the dictionaries when deleting
    if row_num in selected_cannot_days:
        del selected_cannot_days[row_num]
    if row_num in selected_prefer_days:
        del selected_prefer_days[row_num]
    if row_num in selected_manual_days:
        del selected_manual_days[row_num]

    # Find and destroy all widgets in this row
    for row_widgets in worker_rows:
        if row_widgets['row_num'] == row_num:
            # Destroy each widget
            for key, widget in row_widgets.items():
                if key != 'row_num' and widget.winfo_exists():
                    widget.destroy()
            # Remove from worker_rows list
            worker_rows.remove(row_widgets)
            break

    # Remove the worker from workers_list if it exists
    workers_list = [worker for worker in workers_list if worker.get("worker_row_number") != row_num]
    error_label.config(text=f"Worker '{worker_name}' deleted and all their shifts unassigned!")

def pulp_settings():
    popup = Toplevel(root)
    popup.title("PuLP Settings")

    Label(popup, text="Points for shifts filled:").grid(row=0, column=0, sticky="w")
    filled_entry = Entry(popup)
    filled_entry.insert(0, str(points_filled))
    filled_entry.grid(row=0, column=1)

    Label(popup, text="Points for preferred shifts:").grid(row=1, column=0, sticky="w")
    preferred_entry = Entry(popup)
    preferred_entry.insert(0, str(points_preferred))
    preferred_entry.grid(row=1, column=1)

    Label(popup, text="Points for bad spacing days:").grid(row=2, column=0, sticky="w")
    spacing_entry = Entry(popup)
    spacing_entry.insert(0, str(points_spacing))
    spacing_entry.grid(row=2, column=1)

    Label(popup, text="Days apart for spacing penalty:").grid(row=3, column=0, sticky="w")
    spacing_days_entry = Entry(popup)
    spacing_days_entry.insert(0, str(spacing_days_threshold))
    spacing_days_entry.grid(row=3, column=1)

    Label(popup, text="Points for 24-hour shifts:").grid(row=4, column=0, sticky="w")
    hr24_entry = Entry(popup)
    hr24_entry.insert(0, str(points_24hr))
    hr24_entry.grid(row=4, column=1)

    Label(popup, text="Enforce: No Day → Day").grid(row=5, column=0, sticky="w")
    adj_days_var = IntVar(value=1 if enforce_no_adj_days else 0)
    Checkbutton(popup, variable=adj_days_var).grid(row=5, column=1)

    Label(popup, text="Enforce: No Night → Night").grid(row=6, column=0, sticky="w")
    adj_nights_var = IntVar(value=1 if enforce_no_adj_nights else 0)
    Checkbutton(popup, variable=adj_nights_var).grid(row=6, column=1)

    Label(popup, text="Include Mon-Fri day shifts").grid(row=7, column=0, sticky="w")
    include_weekday_var = IntVar(value=1 if include_weekday_days else 0)
    Checkbutton(popup, variable=include_weekday_var).grid(row=7, column=1)

    def save_settings():
        global points_filled, points_preferred, points_spacing, spacing_days_threshold, points_24hr
        global enforce_no_adj_days, enforce_no_adj_nights, include_weekday_days
        try:
            points_filled = int(filled_entry.get())
            points_preferred = int(preferred_entry.get())
            points_spacing = int(spacing_entry.get())
            spacing_days_threshold = int(spacing_days_entry.get())
            points_24hr = int(hr24_entry.get())

            enforce_no_adj_days   = bool(adj_days_var.get())
            enforce_no_adj_nights = bool(adj_nights_var.get())
            include_weekday_days  = bool(include_weekday_var.get())

            error_label.config(text="PuLP settings updated successfully!")

            popup.destroy()
        except ValueError:
            error_label.config(text="Error: All values must be integers.")

    Button(popup, text="Save Settings", command=save_settings).grid(row=8, column=0, columnspan=2)

def save_preferences():
    if year is None:
        error_label.config(text="Year not set.")
        return
    file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json"), ("All files", "*.*")])
    if file_path:
        # Clean up the dictionaries before saving
        # Get a set of row numbers that actually have saved workers
        saved_row_numbers = set()  # Like a basket to collect valid table numbers
        for worker in workers_list:  # Loop through all saved workers
            saved_row_numbers.add(worker["worker_row_number"])  # Add their table number
        
        # Filter the dictionaries to only keep entries for saved workers
        # This is like throwing away notes from empty tables
        cleaned_cannot = {row_num: days for row_num, days in selected_cannot_days.items() if row_num in saved_row_numbers}
        cleaned_prefer = {row_num: days for row_num, days in selected_prefer_days.items() if row_num in saved_row_numbers}
        cleaned_manual = {row_num: days for row_num, days in selected_manual_days.items() if row_num in saved_row_numbers}
        
        # Build the data to save (using cleaned versions)
        data = {"year": year}
        if month is not None:
            data["month"] = month
        data["holiday_days"] = holiday_days
        data["shifts_list"] = shifts_list
        data["workers_list"] = workers_list
        data["selected_cannot_days"] = cleaned_cannot  # Use cleaned version
        data["selected_prefer_days"] = cleaned_prefer  # Use cleaned version
        data["selected_manual_days"] = cleaned_manual  # Use cleaned version
        
        with open(file_path, 'w') as f:
            json.dump(data, f)
        error_label.config(text="Preferences saved.")

def load_preferences():
    global selected_cannot_days, selected_prefer_days, selected_manual_days
    file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json"), ("All files", "*.*")])
    if file_path:
        with open(file_path, 'r') as f:
            data = json.load(f)
        if "year" in data:
            global year
            year = data["year"]
            current_year_label.config(text="Current Year: " + str(year))
            year_entry.delete(0, END)
            year_entry.insert(0, str(year))
        if "month" in data:
            global month
            month = data["month"]
            month_entry.delete(0, END)
            month_entry.insert(0, str(month))
            month_name_list = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
            month_name = month_name_list[month-1]
            current_month_label.config(text="Current Month: " + month_name)
            # Since month is loaded, call save_month to set days_list etc.
            save_month()
        if "holiday_days" in data:
            global holiday_days
            holiday_days = data["holiday_days"]
            holiday_entry.delete(0, END)
            if holiday_days:
                holiday_entry.insert(0, ", ".join(map(str, holiday_days)))
            holidays_label.config(text="Holiday List: " + str(holiday_days))
        if "shifts_list" in data:
            global shifts_list
            shifts_list = data["shifts_list"]
        if "workers_list" in data:
            global workers_list
            workers_list = data["workers_list"]
            # Populate selected_cannot_days from workers_list if not in data
            if "selected_cannot_days" not in data:
                selected_cannot_days = {worker["worker_row_number"]: worker.get("cannot_work", []) for worker in workers_list}
            # Clear existing worker rows
            global worker_rows, worker_row_number
            for row_widgets in worker_rows:
                for key, widget in row_widgets.items():
                    if key != 'row_num' and widget.winfo_exists():
                        widget.destroy()
            worker_rows = []
            worker_row_number = 1
            # Now add rows for loaded workers
            for worker in workers_list:
                row_num = worker["worker_row_number"]
                worker_row_number = row_num
                add_worker_row()
            # Now populate the entries
            for worker in workers_list:
                row_num = worker["worker_row_number"]
                for row_widgets in worker_rows:
                    if row_widgets['row_num'] == row_num:
                        row_widgets['name_entry'].delete(0, END)
                        row_widgets['name_entry'].insert(0, worker["name"])
                        row_widgets['range_entry'].delete(0, END)
                        row_widgets['range_entry'].insert(0, f"{worker['shifts_to_fill'][0]}-{worker['shifts_to_fill'][1]}")
                        row_widgets['max_weekends_entry'].delete(0, END)
                        row_widgets['max_weekends_entry'].insert(0, str(worker["max_weekends"]))
                        row_widgets['max_24hr_entry'].delete(0, END)
                        row_widgets['max_24hr_entry'].insert(0, str(worker["max_24hr"]))
                        break
            # Update worker_row_number to max +1
            if workers_list:
                max_row = max(worker["worker_row_number"] for worker in workers_list)
                worker_row_number = max_row + 1
            else:
                worker_row_number = 1
        if "selected_cannot_days" in data:
            selected_cannot_days = {int(k): v for k, v in data["selected_cannot_days"].items()}
        if "selected_prefer_days" in data:
            selected_prefer_days = {int(k): v for k, v in data["selected_prefer_days"].items()}
        if "selected_manual_days" in data:
            selected_manual_days = {int(k): v for k, v in data["selected_manual_days"].items()}
        # Set worker dicts from loaded selected dicts to ensure consistency
        for worker in workers_list:
            row_num = worker["worker_row_number"]
            worker["cannot_work"] = selected_cannot_days.get(row_num, [])
            worker["prefers"] = selected_prefer_days.get(row_num, [])
            if row_num not in selected_manual_days:
                selected_manual_days[row_num] = []
        error_label.config(text="Preferences loaded.")
        # Update button texts to show selections
        for row_widgets in worker_rows:
            row_num = row_widgets['row_num']
            num_cannot = len(selected_cannot_days.get(row_num, []))
            row_widgets['cannot_button'].config(text=f"Select ({num_cannot})" if num_cannot > 0 else "Select")
            num_prefer = len(selected_prefer_days.get(row_num, []))
            row_widgets['prefer_button'].config(text=f"Select ({num_prefer})" if num_prefer > 0 else "Select")
            num_manual = len(selected_manual_days.get(row_num, []))
            row_widgets['manual_button'].config(text=f"Select ({num_manual})" if num_manual > 0 else "Select")
    print(f"Debugging. Year: {year}, month {month}, Holiday list: {holiday_days}, Shift list length {len(shifts_list)}")
    for worker in workers_list:
        print(worker)
    print(selected_cannot_days)

def load_xlsx_preferences(): 
    """
    Load worker preferences from an Excel (.xlsx) file.
    
    Expected format:
    - Row 1: Headers with "Name" in column A, day numbers (1, 2, 3...) in columns B onwards
    - Row 2+: Worker names in column A, colored cells for preferences
    - RED cells = Cannot work that day (both Day & Night shifts)
    - GREEN cells = Prefer to work that day (both Day & Night shifts)
    """
    global workers_list, worker_rows, worker_row_number
    global selected_cannot_days, selected_prefer_days, selected_manual_days

    # IMPORTANT: First check if days_list exists (year and month must be set)
    try:
        if not days_list:
            error_label.config(text="Error: Please set year and month first!")
            return
    except NameError:
        error_label.config(text="Error: Please set year and month first!")
        return

    # Open file dialog
    file_path = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
    )
    if not file_path:
        return  # User clicked Cancel

    try:
        # data_only=False → keep formulas and formatting (needed to read cell colors)
        # data_only=True would only give us the VALUES, losing all color information
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        # Try to find a suitable sheet
        sheet = None
        
        # If only one sheet, just use it
        if len(wb.sheetnames) == 1:
            sheet = wb[wb.sheetnames[0]]
        else:
            # Try to find sheet by name (supports multiple languages)
            for sname in wb.sheetnames:
                sname_lower = sname.lower()
                if any(keyword in sname_lower for keyword in 
                       ["feuille", "sheet", "preferences", "rota", "workers", "staff"]):
                    sheet = wb[sname]
                    break
        
        # Fallback: use active sheet
        if sheet is None:
            sheet = wb.active

        # Find header row and the "Name" column
        HEADER_ROW = 1
        name_col = None
        
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=HEADER_ROW, column=col).value
            if val and str(val).strip().lower() == "name":
                name_col = col
                break

        if name_col is None:
            error_label.config(text="Error: Could not find a column with header 'Name'")
            return

        # IMPROVED: Build a mapping of day numbers to column numbers
        # This verifies that column B actually corresponds to day 1, etc.
        day_to_col = {}  # {1: 2, 2: 3, 3: 4, ...}
        
        for col in range(1, sheet.max_column + 1):
            header_val = sheet.cell(row=HEADER_ROW, column=col).value
            if header_val is not None:
                try:
                    # Convert to int (handles both 1 and 1.0)
                    day_num = int(float(header_val))
                    # Only accept valid day numbers for this month
                    if 1 <= day_num <= len(days_list):
                        day_to_col[day_num] = col
                except (ValueError, TypeError):
                    pass  # Not a number, skip

        if not day_to_col:
            error_label.config(text="Error: Could not find day number columns (1, 2, 3...)")
            return

        # Clear old data completely
        for row_widgets in worker_rows[:]:  # [:] creates a copy to avoid modification during iteration
            for key, widget in row_widgets.items():
                if key != 'row_num' and widget.winfo_exists():
                    widget.destroy()  # Remove widget from GUI
        
        worker_rows.clear()
        worker_row_number = 1
        workers_list.clear()
        selected_cannot_days.clear()
        selected_prefer_days.clear()
        selected_manual_days.clear()

        loaded_count = 0
        FIRST_DATA_ROW = 2
        total_rows = sheet.max_row - FIRST_DATA_ROW + 1

        # Loop through each worker row
        for row_idx, r in enumerate(range(FIRST_DATA_ROW, sheet.max_row + 1), 1):
            name_cell = sheet.cell(row=r, column=name_col)
            name_val = name_cell.value
            
            # Skip empty rows
            if not name_val:
                continue
            
            name = str(name_val).strip()
            
            # Skip if empty after stripping, or if no letters (probably not a name)
            if not name or not any(c.isalpha() for c in name):
                continue

            # Collect cannot/prefer days by checking cell colors
            cannot_list = []
            prefer_list = []

            # Loop through each day that exists in this month
            for day, col in day_to_col.items():
                cell = sheet.cell(row=r, column=col)
                
                # Check if cell has a solid background color
                if cell.fill and cell.fill.fill_type == 'solid':
                    color = cell.fill.start_color.rgb
                    
                    # CRITICAL FIX: Check if color is None
                    if color is None:
                        continue
                    
                    # Remove alpha channel if present → keep only RRGGBB part
                    color_hex = color.upper()[-6:]  # Last 6 chars = RRGGBB
                    
                    # Check if red (cannot work)
                    if is_red_color(color_hex):
                        cannot_list.append(f"Day {day}")
                        cannot_list.append(f"Night {day}")
                    
                    # Check if green (prefer to work)
                    elif is_green_color(color_hex):
                        prefer_list.append(f"Day {day}")
                        prefer_list.append(f"Night {day}")

            # Create the worker dictionary with default values
            worker_dict = {
                "name": name,
                "shifts_to_fill": [1, 4],  # Default: 0 min, 100 max
                "cannot_work": cannot_list,
                "prefers": prefer_list,
                "max_weekends": 100,
                "max_24hr": 100,
                "worker_row_number": worker_row_number
            }
            workers_list.append(worker_dict)

            # IMPORTANT: Save current row number BEFORE calling add_worker_row
            current_row_num = worker_row_number
            
            # Create GUI row (this will increment worker_row_number internally)
            add_worker_row()
            
            # CRITICAL FIX: Do NOT increment worker_row_number here!
            # add_worker_row() already does it, so incrementing again causes double-increment bug
            # REMOVED: worker_row_number += 1

            # Fill in the values in the GUI
            for row_widgets in worker_rows:
                if row_widgets['row_num'] == current_row_num:
                    # Fill name
                    row_widgets['name_entry'].delete(0, END)
                    row_widgets['name_entry'].insert(0, name)
                    
                    # Fill shift range (default 0-100)
                    row_widgets['range_entry'].delete(0, END)
                    row_widgets['range_entry'].insert(0, "1-4")
                    
                    # Fill max weekends
                    row_widgets['max_weekends_entry'].delete(0, END)
                    row_widgets['max_weekends_entry'].insert(0, "100")
                    
                    # Fill max 24hr
                    row_widgets['max_24hr_entry'].delete(0, END)
                    row_widgets['max_24hr_entry'].insert(0, "100")

                    # Update button labels to show selection counts
                    num_cannot = len(cannot_list)
                    num_prefer = len(prefer_list)
                    
                    row_widgets['cannot_button'].config(
                        text=f"Select ({num_cannot})" if num_cannot > 0 else "Select"
                    )
                    row_widgets['prefer_button'].config(
                        text=f"Select ({num_prefer})" if num_prefer > 0 else "Select"
                    )
                    row_widgets['manual_button'].config(text="Select")
                    break

            # Store selections for the popup windows
            selected_cannot_days[current_row_num] = cannot_list
            selected_prefer_days[current_row_num] = prefer_list
            selected_manual_days[current_row_num] = []

            loaded_count += 1
            
            # Show progress every 5 workers (helps with large files)
            if loaded_count % 5 == 0 or row_idx == total_rows:
                error_label.config(text=f"Loading... {loaded_count} worker(s)")
                root.update()  # Force GUI to refresh

        # Final success message
        error_label.config(text=f"✓ Loaded {loaded_count} worker(s) — red = cannot, green = prefer")

    except Exception as e:
        error_label.config(text=f"Error reading file: {str(e)}")
        print("Detailed error:", e)  # Print to console for debugging


# Helper functions for color detection
def is_red_color(color_hex):
    """
    Check if a color hex code (RRGGBB) represents red.
    Returns True for pure red and close variations.
    
    Args:
        color_hex: 6-character hex string like "FF0000"
    
    Returns:
        bool: True if the color is red-ish
    """
    try:
        r = int(color_hex[0:2], 16)  # Red channel (0-255)
        g = int(color_hex[2:4], 16)  # Green channel (0-255)
        b = int(color_hex[4:6], 16)  # Blue channel (0-255)
        
        # Red if: red channel high (≥250) and green/blue low (≤5)
        return r >= 250 and g <= 5 and b <= 5
    except (ValueError, IndexError):
        return False


def is_green_color(color_hex):
    """
    Check if a color hex code (RRGGBB) represents green.
    Returns True for pure green and close variations.
    
    Args:
        color_hex: 6-character hex string like "00FF00"
    
    Returns:
        bool: True if the color is green-ish
    """
    try:
        r = int(color_hex[0:2], 16)  # Red channel (0-255)
        g = int(color_hex[2:4], 16)  # Green channel (0-255)
        b = int(color_hex[4:6], 16)  # Blue channel (0-255)
        
        # Green if: green channel high (≥250) and red/blue low (≤5)
        return r <= 5 and g >= 250 and b <= 5
    except (ValueError, IndexError):
        return False

def extract_day_from_shift_name(shift_name):
    """
    Extract day number from shift name.
    Examples: "Day 5 Cardiology" -> 5, "Night 12 Internal Medicine" -> 12
    """
    parts = shift_name.split()
    return int(parts[1])  # Day number is always the second part

def extract_unit_from_shift_name(shift_name):
    """
    Extract unit name from shift name.
    Examples: "Day 5 Cardiology" -> "Cardiology"
    """
    parts = shift_name.split()
    return " ".join(parts[2:])  # Everything after day number is the unit

# ----------------------------------------------------------------------------
# PuLP Solve: solver is in separate file
# ----------------------------------------------------------------------------

def create_rota():
    """
    This is the NEW create_rota function in main.py.
    It's much simpler - it just:
    1. Collects all the data from the GUI
    2. Packages it up
    3. Sends it to solver.py
    4. Gets the results back
    5. Shows the results to the user
    """
    
    # ========================================================================
    # PART A: Gather all the settings into a dictionary
    # ========================================================================
    # Think of this like packing a suitcase before a trip
    
    settings = {
        "points_filled": points_filled,
        "points_preferred": points_preferred,
        "points_spacing": points_spacing,
        "spacing_days_threshold": spacing_days_threshold,
        "points_24hr": points_24hr,
        "enforce_no_adj_nights": enforce_no_adj_nights,
        "enforce_no_adj_days": enforce_no_adj_days
    }
    
    # ========================================================================
    # PART B: Call the solver
    # ========================================================================
    # This is where the magic happens!
    # We send our data to solver.py and get the results back
    
    # The imported create_rota function from solver.py
    assignments, summary = solve_rota(shifts_list, workers_list, units_list, settings)
    
    # ========================================================================
    # PART C: Update the shifts_list with the new assignments
    # ========================================================================
    # The solver returns a dictionary like {"Day 1": "John", "Night 1": "Sarah"}
    # We need to update our shifts_list to match
    
    # for shift in shifts_list:                                     IF THIS IS TURNED ON, then if trying to run create_rota again will run on full rota, even if changed parameters.
    # shift["assigned_worker"] = assignments.get(shift["name"])     IF THIS IS TURNED OFF, only the manually assigned workers stay assigned, the rest are left unassigned and can run create_rota again, with possibly new results
    
    # ========================================================================
    # PART D: Print results to console (for debugging)
    # ========================================================================
    print("Final Rota:")
    for shift_name, worker in assignments.items():
        print(f"{shift_name}: {worker if worker else 'Unassigned'}")
    
    # ========================================================================
    # PART E: Check if solver failed
    # ========================================================================
    if summary["status"] == "Infeasible":
        error_label.config(text="ERROR: Impossible to create rota with current rules! Check shift ranges, max weekends, max 24hr shifts.")
        return  # Stop here - don't show popup
    
    if summary["status"] == "Nothing to assign":
        error_label.config(text="No empty shifts or no workers – nothing to assign.")
        return  # Stop here
    
    # ========================================================================
    # PART F: Create popup window to show results
    # ========================================================================
    popup = Toplevel(root)
    popup.title("Rota Results")
    
    # Create a Text widget (like a mini text editor)
    text_widget = Text(popup, wrap="none", width=60, height=30)
    text_widget.pack(side="left", fill="both", expand=True)
    
    # Create scrollbar
    scrollbar = Scrollbar(popup, orient="vertical", command=text_widget.yview)
    scrollbar.pack(side="right", fill="y")
    text_widget.config(yscrollcommand=scrollbar.set)
    
    # Mouse wheel scrolling
    def on_popup_mousewheel(event):
        text_widget.yview_scroll(int(-1 * (event.delta / 120)), "units")
    
    popup.bind("<MouseWheel>", on_popup_mousewheel)
    text_widget.bind("<MouseWheel>", on_popup_mousewheel)
    
    # Add title
    text_widget.insert("end", "Final Rota (Multi-Unit)\n", "title")
    text_widget.insert("end", "=" * 70 + "\n\n", "separator")
    
    # Group assignments by unit
    assignments_by_unit = {}
    for shift_name, worker in assignments.items():
        if worker is None:
            worker = "Unassigned"
        
        # Extract unit from shift name
        unit = extract_unit_from_shift_name(shift_name)
        
        if unit not in assignments_by_unit:
            assignments_by_unit[unit] = {}
        
        # Extract day number and type
        day = extract_day_from_shift_name(shift_name)
        shift_type = shift_name.split()[0]  # "Day" or "Night"
        
        if day not in assignments_by_unit[unit]:
            assignments_by_unit[unit][day] = {"Day": "No shift", "Night": "No shift"}
        
        assignments_by_unit[unit][day][shift_type] = worker
    
    # Display results grouped by unit
    for unit in sorted(assignments_by_unit.keys()):
        text_widget.insert("end", f"=== {unit} ===\n", "unit_header")
        text_widget.insert("end", "-" * 70 + "\n", "separator")
        text_widget.insert("end", f"{'Day':<7}{'Day Shift':<30}{'Night Shift':<30}\n", "header")
        text_widget.insert("end", "-" * 70 + "\n", "separator")
        
        for day in sorted(assignments_by_unit[unit].keys()):
            day_worker = assignments_by_unit[unit][day].get("Day", "No shift")
            night_worker = assignments_by_unit[unit][day].get("Night", "No shift")
            
            text_widget.insert("end", f"{day:<7}{day_worker:<30}{night_worker:<30}\n")
        
        text_widget.insert("end", "\n")
    
    # Add summary section
    text_widget.insert("end", "=" * 70 + "\n", "separator")
    text_widget.insert("end", "Summary:\n", "title")
    text_widget.insert("end", "-" * 70 + "\n", "separator")
    text_widget.insert("end", f"Number of preferred shifts assigned: {summary['preferences_count']}\n")
    text_widget.insert("end", f"Number of 24-hour shifts: {summary['twenty_four_count']}\n")
    text_widget.insert("end", f"Number of bad spacing pairs (<{spacing_days_threshold} days apart): {summary['bad_spacing_count']}\n")
    
    # Style the text
    text_widget.tag_config("title", font=("Arial", 12, "bold"))
    text_widget.tag_config("unit_header", font=("Arial", 11, "bold"), foreground="blue")
    text_widget.tag_config("header", font=("Arial", 10, "bold"))
    text_widget.tag_config("separator", foreground="gray")
    
    # Make read-only
    text_widget.config(state="disabled")

# ---------------------------------------------------------------
# Small code area mainly for buttons and error sign at the bottom
# ---------------------------------------------------------------

# Add worker button.
Button(add_worker_button_frame, text="Add Worker", command=add_worker_row, width=10, pady=2).pack() 

# Create rota button.
Button(root, text="Create Rota", command=create_rota).pack(pady=4)  # Button to create a rota.

# Frame for settings, saving and loading buttons.
settings_frame = Frame(root)
settings_frame.pack()

Button(settings_frame, text="PuLP Settings", width=14, command=pulp_settings).pack(side=LEFT, padx=4)
Button(settings_frame, text="Save Preferences", width=14, command=save_preferences).pack(side=LEFT, padx=4)
Button(settings_frame, text="Load", width=14, command=load_preferences).pack(side=LEFT, padx=4)
Button(settings_frame, text="Load .xlsx", width=14, command=load_xlsx_preferences).pack(side=LEFT, padx=4)

# Error label (same).
error_label = Label(root, text="")  # For errors.
error_label.pack()

root.mainloop()  # Start the window – like "go!"
